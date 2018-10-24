# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl import load_workbook
wb = load_workbook(r'C:\excel\2018年7月-应收-其他应收-预付款项汇总表-物流.xlsx',data_only=True)
wb.guess_types = True   #猜测格式类型
ws=wb.active

c13 = ws['C13']
#cell函数下标从1开始
c14 = ws.cell(14,3)
print(c13.value)
print(c14.value)

for i,cell in enumerate( list(ws.rows)[13] ):
    if(isinstance(cell.value,float)):
        print(i,round(cell.value,2))
    else:
        print(i,cell.value)

ws.append([1,23,4])

# Save the file
wb.save(r'C:\excel\2018年7月-应收-其他应收-预付款项汇总表-物流.xlsx')
#注意如果原文件有一些图片或者图标，则保存的时候可能会导致图片丢失

# -*- coding: utf-8 -*-
import os
from openpyxl import Workbook
from openpyxl import load_workbook
import winreg

'''
获取子表名字在汇总表的行号，第三个参数表示是否为关联单位
'''
def get_line_number(self_total_sheet,self_cell_value,self_is_related):
    related_company_total = not_related_company_total = 0
    #表格中关联单位是否在前面
    related_in_front = True
    for cell in list(self_total_sheet.columns)[1] :
        #一定要先判断非关联再判断关联，因为如果包含非关联肯定包含关联
        if isinstance(cell.value,str) and "非关联单位小计" in cell.value:
            not_related_company_total = cell.row
        elif isinstance(cell.value,str) and "关联单位小计" in cell.value:
            related_company_total= cell.row

    if(related_company_total < not_related_company_total):
        related_in_front = True
    else:
        related_in_front = False

    #是否需要跳过
    need_skip = False
    if self_is_related:
        if related_in_front:
            need_skip = False
        else:
            need_skip = True
    else:
        if related_in_front:
            need_skip = True
        else:
            need_skip = False

    #这里的i下标是0开始的
    for i,cell in enumerate(list(self_total_sheet.columns)[1]) :
        #print(i,cell.value)
        if cell.value == self_cell_value:
            if need_skip:
                need_skip = False
                continue
            else:
                return i+1;
    return 0


#从子表读入数据
def get_need_data_in_child(self_child_sheet,self_line_num):
    self_result = []
    #2018年期初余额
    cell_value_num_3 = round( self_child_sheet.cell(self_line_num, 3).value,2 )
    print( "2018年期初余额:", cell_value_num_3)
    self_result.append(cell_value_num_3)
    #本期增加
    print("本期增加")
    #计算一年内余额
    relate_cell_value_num_4 = round( self_child_sheet.cell(self_line_num, 4).value,2 )
    relate_cell_value_num_5 = round( self_child_sheet.cell(self_line_num, 5).value,2 )
    relate_cell_value_num_6 = round( self_child_sheet.cell(self_line_num, 6).value,2 )
    relate_cell_value_num_increase_year = round( relate_cell_value_num_4 + relate_cell_value_num_5 + relate_cell_value_num_6 , 2)
    print("1年内:" + str(relate_cell_value_num_increase_year))
    self_result.append(relate_cell_value_num_increase_year)
    #计算1-2年余额
    relate_cell_value_num_7 = round( self_child_sheet.cell(self_line_num, 7).value,2 )
    print("1-2年:" + str(relate_cell_value_num_7))
    self_result.append(relate_cell_value_num_7)
    #计算2-3年余额
    relate_cell_value_num_8 = round( self_child_sheet.cell(self_line_num, 8).value,2 )
    print("2-3年:" + str(relate_cell_value_num_8))
    self_result.append(relate_cell_value_num_8)
    #计算3年以上余额
    relate_cell_value_num_9 = round( self_child_sheet.cell(self_line_num, 9).value,2 )
    print("3年以上:" + str(relate_cell_value_num_9))
    self_result.append(relate_cell_value_num_9)
    #本期减少
    print("本期减少")
    #计算一年内余额
    relate_cell_value_num_10 = round( self_child_sheet.cell(self_line_num, 10).value,2 )
    relate_cell_value_num_11 = round( self_child_sheet.cell(self_line_num, 11).value,2 )
    relate_cell_value_num_12 = round( self_child_sheet.cell(self_line_num, 12).value,2 )
    relate_cell_value_num_decrease_year = round( relate_cell_value_num_10 + relate_cell_value_num_11 + relate_cell_value_num_12 , 2)
    print("1年内:" + str(relate_cell_value_num_decrease_year))
    self_result.append(relate_cell_value_num_decrease_year)
    #计算1-2年余额
    relate_cell_value_num_13 = round( self_child_sheet.cell(self_line_num, 13).value,2 )
    print("1-2年:" + str(relate_cell_value_num_13))
    self_result.append(relate_cell_value_num_13)
    #计算2-3年余额
    relate_cell_value_num_14 = round( self_child_sheet.cell(self_line_num, 14).value,2 )
    print("2-3年:" + str(relate_cell_value_num_14))
    self_result.append(relate_cell_value_num_14)
    #计算3年以上余额
    relate_cell_value_num_15 = round( self_child_sheet.cell(self_line_num, 15).value,2 )
    print("3年以上:" + str(relate_cell_value_num_15))
    self_result.append(relate_cell_value_num_15)
    #2018年6月末余额
    relate_cell_value_num_16 = round( self_child_sheet.cell(self_line_num, 16).value,2 )
    print( "2018年6月末余额:", relate_cell_value_num_16)
    self_result.append(relate_cell_value_num_16)
    #计算一年内余额
    relate_cell_value_num_17 = round( self_child_sheet.cell(self_line_num, 17).value,2 )
    relate_cell_value_num_18 = round( self_child_sheet.cell(self_line_num, 18).value,2 )
    relate_cell_value_num_19 = round( self_child_sheet.cell(self_line_num, 19).value,2 )
    relate_cell_value_num_year = round( relate_cell_value_num_17 + relate_cell_value_num_18 + relate_cell_value_num_19 , 2)
    print("1年内:" + str(relate_cell_value_num_year))
    self_result.append(relate_cell_value_num_year)
    #计算1-2年余额
    relate_cell_value_num_20 = round( self_child_sheet.cell(self_line_num, 20).value,2 )
    print("1-2年:" + str(relate_cell_value_num_20))
    self_result.append(relate_cell_value_num_20)
    #计算2-3年余额
    relate_cell_value_num_21 = round( self_child_sheet.cell(self_line_num, 21).value,2 )
    print("2-3年:" + str(relate_cell_value_num_21))
    self_result.append(relate_cell_value_num_21)
    #计算3年以上余额
    relate_cell_value_num_22 = round( self_child_sheet.cell(self_line_num, 22).value,2 )
    print("3年以上:" + str(relate_cell_value_num_22))
    self_result.append(relate_cell_value_num_22)
    return self_result


#向汇总表写入数据
def write_need_data_to_total(self_total_sheet,self_line_num,result):
    self_total_sheet.cell(self_line_num, 3).value = result[0]
    self_total_sheet.cell(self_line_num, 4).value = result[1]
    self_total_sheet.cell(self_line_num, 5).value = result[2]
    self_total_sheet.cell(self_line_num, 6).value = result[3]
    self_total_sheet.cell(self_line_num, 7).value = result[4]
    self_total_sheet.cell(self_line_num, 8).value = result[5]
    self_total_sheet.cell(self_line_num, 9).value = result[6]
    self_total_sheet.cell(self_line_num, 10).value = result[7]
    self_total_sheet.cell(self_line_num, 11).value = result[8]
    self_total_sheet.cell(self_line_num, 12).value = result[9]
    self_total_sheet.cell(self_line_num, 13).value = result[10]
    self_total_sheet.cell(self_line_num, 14).value = result[11]
    self_total_sheet.cell(self_line_num, 15).value = result[12]
    self_total_sheet.cell(self_line_num, 16).value = result[13]

def get_key_name(self_child_file_location):
    suffix = os.path.splitext(self_child_file_location)[-1]
    #去掉后缀名
    last = self_child_file_location.split("-")[-1].replace(suffix,"")
    return last

#获取桌面路径
def get_desktop():
    key = winreg.OpenKey(winreg.HKEY_CURRENT_USER,r'Software\Microsoft\Windows\CurrentVersion\Explorer\Shell Folders')#利用系统的链表
    return winreg.QueryValueEx(key, "Desktop")[0] #返回的是Unicode类型数据


def convert(self_child_file_location,self_total_file_location):
    child_book = load_workbook(self_child_file_location, data_only=True)
    child_book.guess_types = True   #猜测格式类型
    child_sheet=child_book.active

    total_book = load_workbook(self_total_file_location, data_only=True)
    total_book.guess_types = True   #猜测格式类型
    total_sheet=total_book.active



    # Save the file
    child_book.save(self_child_file_location)
    total_book.save(self_total_file_location)
    #注意如果原文件有一些图片或者图标，则保存的时候可能会导致图片丢失


#真正的转换操作
convert(r"C:/Users/lenovo/Desktop/党费/能力平台事业部党支部支部2018年8月党费明细.xlsx",
        r"C:/Users/lenovo/Desktop/党费/能力平台事业部党支部2018年1-7月党费明细台账.xlsx")
